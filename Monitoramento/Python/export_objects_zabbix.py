#!/usr/bin/python3

"""
Nome do Script: /usr/lib/zabbix/externalscripts/export_objects_zabbix.py
Sistema: Zabbix
Criado por: Wander Maia da Silva
Data da Criação: 19/10/2020
Descrição: Script para realizar o export dos objetos (hosts, triggers, itens, usuários,etc) configurados no Zabbix
Data da última modificação: 29/10/2020
Modificação: Acrescentado parâmetro para exibir as expressões das triggers.
Entrada: Dados para acesso à API do Zabbix
Saida: Planilha excel contendo os dados dos objetos do ambiente
Linha do Plugin: ./check_objects_zabbix.py -s 'localhost' -u 'Automation' -p 'zabbix'

"""


"""
Bibliotecas necessáias para o menu de opções, API do Zabbix e operações com data/hora
"""
import argparse
from zabbix_api import ZabbixAPI
from datetime import datetime
from openpyxl import Workbook



# Criando o parser para validar as opções do script e apresentar a ajuda
parser = argparse.ArgumentParser(description="Script para realizar o export dos objetos (hosts, triggers, itens, usuários,etc) configurados no Zabbix")

# Definindo os argumentos
parser.add_argument("-s", "--server", action = 'store', dest = 'serverName', type=str, required = True,
                    help="Server Zabbix para conexão. Exemplos: 192.168.56.195, localhost ou Zabbix.domain.local")
parser.add_argument("-u", "--user", action = 'store', dest = 'userName', type=str, required = True,
                    help="Nome do usuário para conexão na API do Zabbix")
parser.add_argument("-p", "--password", action = 'store', dest = 'userPassword', type=str, required = True,
                    help="Senha do usuario para conexão na API do Zabbix")

# Passandos os parâmetros para a variável adequada
args = parser.parse_args()


# Montando a URL da API e segregando os valores de acesso
zabbix_server = "http://{}/zabbix".format(args.serverName)
username = args.userName
password = args.userPassword


# Instanciando a conexao com a API
conexao = ZabbixAPI(server = zabbix_server)
conexao.login(username, password)


"""
Realizando a coleta dos dados dos hosts. Também está sendo criada uma lista contendo os IDs de todos os hosts cadastrados a partir da coleta realizada pela API.
Esta lista será utilizada em filtros nas coletas posteriores.
"""
hosts = conexao.host.get({"output": "extend", "sortfield": "name"})
listaIdHosts = []
for host in hosts:
    listaIdHosts.append("{}".format(host['hostid']))



"""
Realizando a coleta dos dados dos demais dados necessários. São coletados os itens, triggers e usuários.
Para facilitar o preenchimento dos dados das triggers, foi utilizada a lista de IDs de hosts.
Caso esta lista não fosse utilizada, também seriam retornados pela API as triggers cadastradas nos modelos.
"""
items = conexao.item.get({"output": "extend", "monitored": "true", "sortfield": "itemid"})
triggers = conexao.trigger.get({"output": "extend", "sortfield": "triggerid", "hostids" : listaIdHosts ,  "selectHosts": ["hostid", "host"], "expandExpression": "1" })
users = conexao.user.get({"output": "extend"})


"""
Criando o workbook (arquivo excel) que conterá as planilhas com os dados coletados.
A planilha inicial, que é gerada automaticamente, foi ativada e ajustado o nome para que seja utilizada para os dados dos hosts.
"""
arquivoExcel = Workbook()
planilhaHosts = arquivoExcel.active
planilhaHosts.title = "Hosts"

"""
Criando as demais planilhas que serão utilizadas para os dados.
Estão sendo ciradas as de items, triggers e users. Todas estão sendo criadas no mesmo arquivo
"""
planilhaItens= arquivoExcel.create_sheet("Itens")
planilhaTriggers= arquivoExcel.create_sheet("Triggers")
planilhaUsers= arquivoExcel.create_sheet("Users")




"""
Função utilizada para verificar o IP do host. Ela recebe o ID do Host e realiza uma validação pelo ID do host e retorna
o IP da primeira interface do host.
"""
def verificaIpHost(idHost):
    
    # Coleta as interfaces para o hostid informado no parâmetro.
    hostInterfaces = conexao.hostinterface.get({"output": "extend", "hostids": "{}".format(idHost)})
    return hostInterfaces[0]['ip']


"""
Função para verificar o nome do host que contém o ID informado.
"""
def verificaHostname(idHost):
    
    # Loop que verifica e retorna o nome do host com a chave informada.
    for host in hosts:
        if idHost in host.values():
            return host['name']





"""
Inserindo os dados dos Hosts. Para inserir os dados dos hosts, primeiramente inserimos a primeira linha como títulos das colunas.
Em seguida, é executado o loop para a inserção de dados de todos os hosts.
"""
titulosPlanilhaHosts = ("hostid", "host", "name", "IP", "status", "maintenance_status")
planilhaHosts.append(titulosPlanilhaHosts)
for host in hosts:
    dadosHost = (host['hostid'],host['host'],host['name'],verificaIpHost(host['hostid']),host['status'], host['maintenance_status'])
    planilhaHosts.append(dadosHost)



"""
Inserindo os dados dos itens. Para inserir os dados dos itens, primeiramente inserimos a primeira linha como títulos das colunas.
Em seguida, é executado o loop para a inserção de dados de todos os itens.
"""
titulosPlanilhaItens = ("hostname", "itemid", "name", "state", "key", "delay","history","trends","state","status")
planilhaItens.append(titulosPlanilhaItens)
for item in items:
    hostname = verificaHostname(item['hostid'])
    dadosItem = (hostname,item['itemid'],item['name'],item['state'],item['key_'],item['delay'],item['history'],item['trends'],item['state'],item['status'])
    planilhaItens.append(dadosItem)


"""
Inserindo os dados das Triggers. Para inserir os dados dos triggers, primeiramente inserimos a primeira linha como títulos das colunas.
Em seguida, é executado o loop para a inserção de dados de todas as triggers.
"""
titulosPlanilhaTriggers = ("host","triggerid", "description", "expression", "status", "priority", "comments","state","recovery_mode","recovery_expression","manual_close")
planilhaTriggers.append(titulosPlanilhaTriggers)
for trigger in triggers:
    dadosTrigger = (trigger['hosts'][0]['host'],trigger['triggerid'],trigger['description'],trigger['expression'],trigger['status'],trigger['priority'],trigger['comments'],trigger['state'],trigger['recovery_mode'],trigger['recovery_expression'],trigger['manual_close'])
    planilhaTriggers.append(dadosTrigger)



"""
Inserindo os dados dos Usuários. Para inserir os dados dos usuários, primeiramente inserimos a primeira linha como títulos das colunas.
Em seguida, é executado o loop para a inserção dos dados de todos as usuários.
"""
titulosPlanilhaUsers = ("userid", "username", "name", "surname", "autologin", "autologout","language","roleid","theme")
planilhaUsers.append(titulosPlanilhaUsers)
for user in users:
    dadosUser = (user['userid'],user['username'],user['name'],user['surname'],user['autologin'],user['autologout'],user['lang'],user['roleid'],user['theme'])
    planilhaUsers.append(dadosUser)



"""
Gerando o nome do arquivo utilizando o horário da coleta e salvando o arquivo contendo os dados coletados.
"""
dataHora = datetime.now()
nomePlanilha = "/tmp/zabbix_objects_{}.xlsx".format(dataHora.strftime('%Y%m%d_%H%M'))
arquivoExcel.save(nomePlanilha)
print("Planilha de export de Objetos gerada em: '{}'".format(nomePlanilha))

